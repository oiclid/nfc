"""
Report generator — produces PDF and Excel reports.
All methods return the output file path.
"""
import os
import sqlite3
from datetime import datetime, date
from typing import Optional, List, Dict

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'data', 'reports'
)

HEADER_COLOR  = colors.HexColor('#1A5276')
ACCENT_COLOR  = colors.HexColor('#2980B9')
ALT_ROW_COLOR = colors.HexColor('#EBF5FB')
WHITE         = colors.white
DARK_TEXT     = colors.HexColor('#1C2833')

XL_HEADER_FILL  = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
XL_SUBHDR_FILL  = PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid')
XL_ALT_FILL     = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
XL_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


class ReportGenerator:
    def __init__(self, db_path: str):
        self.db_path = db_path
        os.makedirs(REPORTS_DIR, exist_ok=True)

    # -------------------------------------------------------------------------
    # DB helpers
    # -------------------------------------------------------------------------

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _q(self, query: str, params: tuple = ()) -> List[Dict]:
        conn = self._conn()
        rows = [dict(r) for r in conn.execute(query, params).fetchall()]
        conn.close()
        return rows

    def _setting(self, key: str) -> str:
        rows = self._q(
            "SELECT setting_value FROM system_settings WHERE setting_key=?", (key,)
        )
        return rows[0]['setting_value'] if rows else ''

    def _org(self) -> str:
        return self._setting('organization_name') or 'NFC Cooperative'

    def _currency(self) -> str:
        return self._setting('currency_symbol') or '₦'

    def _fmt(self, amount) -> str:
        return f"{self._currency()}{float(amount or 0):,.2f}"

    def _ts(self) -> str:
        return datetime.now().strftime('%Y%m%d_%H%M%S')

    def _out(self, name: str, ext: str) -> str:
        return os.path.join(REPORTS_DIR, f"{name}_{self._ts()}.{ext}")

    # -------------------------------------------------------------------------
    # PDF helpers
    # -------------------------------------------------------------------------

    def _styles(self):
        s = getSampleStyleSheet()
        s.add(ParagraphStyle('OrgName',   fontName='Helvetica-Bold',  fontSize=16,
                              textColor=HEADER_COLOR,  alignment=TA_CENTER))
        s.add(ParagraphStyle('ReportTitle', fontName='Helvetica-Bold', fontSize=13,
                              textColor=DARK_TEXT,     alignment=TA_CENTER))
        s.add(ParagraphStyle('SubTitle',  fontName='Helvetica',       fontSize=10,
                              textColor=colors.grey,   alignment=TA_CENTER))
        s.add(ParagraphStyle('SectionHdr', fontName='Helvetica-Bold', fontSize=11,
                              textColor=HEADER_COLOR,  spaceBefore=12, spaceAfter=4))
        s.add(ParagraphStyle('TableCell', fontName='Helvetica',       fontSize=9,
                              textColor=DARK_TEXT))
        s.add(ParagraphStyle('TableHdr',  fontName='Helvetica-Bold',  fontSize=9,
                              textColor=WHITE))
        s.add(ParagraphStyle('Footer',    fontName='Helvetica',       fontSize=8,
                              textColor=colors.grey,   alignment=TA_CENTER))
        s.add(ParagraphStyle('Summary',   fontName='Helvetica-Bold',  fontSize=10,
                              textColor=DARK_TEXT))
        return s

    def _pdf_header(self, styles, title: str, subtitle: str = '') -> list:
        items = [
            Paragraph(self._org(), styles['OrgName']),
            Spacer(1, 4),
            Paragraph(title, styles['ReportTitle']),
        ]
        if subtitle:
            items.append(Paragraph(subtitle, styles['SubTitle']))
        items += [
            Paragraph(
                f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
                styles['SubTitle']
            ),
            HRFlowable(width='100%', thickness=1.5, color=HEADER_COLOR),
            Spacer(1, 8),
        ]
        return items

    def _pdf_table_style(self, n_cols: int, header_rows: int = 1) -> TableStyle:
        return TableStyle([
            ('BACKGROUND',   (0, 0),           (-1, header_rows - 1), HEADER_COLOR),
            ('TEXTCOLOR',    (0, 0),           (-1, header_rows - 1), WHITE),
            ('FONTNAME',     (0, 0),           (-1, header_rows - 1), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0),           (-1, header_rows - 1), 9),
            ('ALIGN',        (0, 0),           (-1, header_rows - 1), 'CENTER'),
            ('FONTNAME',     (0, header_rows), (-1, -1),              'Helvetica'),
            ('FONTSIZE',     (0, header_rows), (-1, -1),              8),
            ('ROWBACKGROUNDS', (0, header_rows), (-1, -1),            [WHITE, ALT_ROW_COLOR]),
            ('GRID',         (0, 0),           (-1, -1),              0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN',       (0, 0),           (-1, -1),              'MIDDLE'),
            ('TOPPADDING',   (0, 0),           (-1, -1),              4),
            ('BOTTOMPADDING',(0, 0),           (-1, -1),              4),
            ('LEFTPADDING',  (0, 0),           (-1, -1),              6),
            ('RIGHTPADDING', (0, 0),           (-1, -1),              6),
        ])

    # -------------------------------------------------------------------------
    # Excel helpers
    # -------------------------------------------------------------------------

    def _xl_header_row(self, ws, row: int, headers: list, col_start: int = 1):
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=col_start + i, value=h)
            cell.font      = Font(bold=True, color='FFFFFF', size=10)
            cell.fill      = XL_HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = XL_BORDER

    def _xl_data_row(self, ws, row: int, values: list, alt: bool = False, col_start: int = 1):
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=col_start + i, value=v)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border    = XL_BORDER
            if alt:
                cell.fill = XL_ALT_FILL

    def _xl_title(self, ws, title: str, subtitle: str, n_cols: int):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=self._org())
        c.font      = Font(bold=True, size=14, color='1A5276')
        c.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c = ws.cell(row=2, column=1, value=title)
        c.font      = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
        c = ws.cell(row=3, column=1, value=subtitle)
        c.font      = Font(italic=True, size=10, color='666666')
        c.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        c = ws.cell(row=4, column=1,
                    value=f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
        c.font      = Font(italic=True, size=9, color='888888')
        c.alignment = Alignment(horizontal='center')

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        return 5  # next available row

    # -------------------------------------------------------------------------
    # 1. Members List
    # -------------------------------------------------------------------------

    def members_list_pdf(self, station_id: Optional[str] = None,
                          status: str = 'Active') -> str:
        q = """
            SELECT m.member_id, m.first_name, m.middle_name, m.last_name,
                   m.gender, m.date_joined, m.phone_number, m.grade_level,
                   s.station_name
            FROM members m JOIN stations s ON m.station_id=s.station_id
            WHERE 1=1
        """
        params = []
        if station_id:
            q += " AND m.station_id=?"; params.append(station_id)
        if status == 'Active':
            q += " AND m.is_active=1 AND m.is_deceased=0"
        elif status == 'Inactive':
            q += " AND m.is_active=0 AND m.is_deceased=0"
        elif status == 'Deceased':
            q += " AND m.is_deceased=1"
        q += " ORDER BY m.member_id"
        members = self._q(q, tuple(params))

        out    = self._out('members_list', 'pdf')
        doc    = SimpleDocTemplate(out, pagesize=landscape(A4),
                                   leftMargin=1.5*cm, rightMargin=1.5*cm,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = self._styles()
        story  = self._pdf_header(styles, f"Members List — {status}",
                                   f"Total: {len(members)}")

        headers = ['ID', 'First Name', 'Middle', 'Last Name', 'Gender',
                   'Station', 'Date Joined', 'Phone', 'Grade']
        rows    = [headers]
        for m in members:
            rows.append([
                m['member_id'], m['first_name'], m['middle_name'] or '',
                m['last_name'], m['gender'] or '', m['station_name'],
                m['date_joined'] or '', m['phone_number'] or '',
                m['grade_level'] or ''
            ])

        col_widths = [2.2*cm, 3*cm, 2.5*cm, 3*cm, 1.8*cm,
                      4*cm, 2.5*cm, 3*cm, 2*cm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(self._pdf_table_style(len(headers)))
        story.append(t)
        doc.build(story)
        return out

    def members_list_excel(self, station_id: Optional[str] = None,
                            status: str = 'Active') -> str:
        q = """
            SELECT m.member_id, m.first_name, m.middle_name, m.last_name,
                   m.gender, m.date_joined, m.phone_number, m.employee_id,
                   m.grade_level, m.nok1_name, m.nok1_phone, s.station_name
            FROM members m JOIN stations s ON m.station_id=s.station_id
            WHERE 1=1
        """
        params = []
        if station_id:
            q += " AND m.station_id=?"; params.append(station_id)
        if status == 'Active':
            q += " AND m.is_active=1 AND m.is_deceased=0"
        elif status == 'Inactive':
            q += " AND m.is_active=0 AND m.is_deceased=0"
        elif status == 'Deceased':
            q += " AND m.is_deceased=1"
        q += " ORDER BY m.member_id"
        members = self._q(q, tuple(params))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"

        headers = ['Member ID', 'First Name', 'Middle Name', 'Last Name',
                   'Gender', 'Station', 'Date Joined', 'Phone', 'Employee ID',
                   'Grade Level', 'NOK Name', 'NOK Phone']
        next_row = self._xl_title(ws, f'Members List — {status}',
                                   f'Total: {len(members)}', len(headers))
        self._xl_header_row(ws, next_row, headers)
        next_row += 1

        for i, m in enumerate(members):
            self._xl_data_row(ws, next_row, [
                m['member_id'], m['first_name'], m['middle_name'] or '',
                m['last_name'], m['gender'] or '', m['station_name'],
                m['date_joined'] or '', m['phone_number'] or '',
                m['employee_id'] or '', m['grade_level'] or '',
                m['nok1_name'] or '', m['nok1_phone'] or ''
            ], alt=(i % 2 == 1))
            next_row += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        out = self._out('members_list', 'xlsx')
        wb.save(out)
        return out

    # -------------------------------------------------------------------------
    # 2. Savings Summary
    # -------------------------------------------------------------------------

    def savings_summary_pdf(self, station_id: Optional[str] = None) -> str:
        q = """
            SELECT m.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name,
                   ROUND(SUM(CASE WHEN st.type_code='PREMIUM' THEN sa.current_balance ELSE 0 END),2) AS premium,
                   ROUND(SUM(CASE WHEN st.type_code='TARGET'  THEN sa.current_balance ELSE 0 END),2) AS target,
                   ROUND(SUM(CASE WHEN st.type_code='SHARES'  THEN sa.current_balance ELSE 0 END),2) AS shares,
                   ROUND(SUM(sa.current_balance),2) AS total
            FROM members m
            JOIN stations s ON m.station_id=s.station_id
            JOIN savings_accounts sa ON m.member_id=sa.member_id
            JOIN savings_types st ON sa.savings_type_id=st.savings_type_id
            WHERE m.is_active=1 AND m.is_deceased=0 AND sa.is_active=1
        """
        params = []
        if station_id:
            q += " AND m.station_id=?"; params.append(station_id)
        q += " GROUP BY m.member_id ORDER BY m.member_id"
        rows = self._q(q, tuple(params))

        grand_premium = sum(r['premium'] or 0 for r in rows)
        grand_target  = sum(r['target'] or 0 for r in rows)
        grand_shares  = sum(r['shares'] or 0 for r in rows)
        grand_total   = sum(r['total'] or 0 for r in rows)
        currency      = self._currency()

        out    = self._out('savings_summary', 'pdf')
        doc    = SimpleDocTemplate(out, pagesize=landscape(A4),
                                   leftMargin=1.5*cm, rightMargin=1.5*cm,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = self._styles()
        story  = self._pdf_header(styles, "Savings Summary Report",
                                   f"Members: {len(rows)}  |  "
                                   f"Grand Total: {currency}{grand_total:,.2f}")

        headers = ['Member ID', 'Full Name', 'Station',
                   'Premium', 'Target/Special', 'Shares', 'Total']
        table_rows = [headers]
        for r in rows:
            table_rows.append([
                r['member_id'], r['full_name'], r['station_name'],
                f"{currency}{r['premium'] or 0:,.2f}",
                f"{currency}{r['target'] or 0:,.2f}",
                f"{currency}{r['shares'] or 0:,.2f}",
                f"{currency}{r['total'] or 0:,.2f}",
            ])
        # Totals row
        table_rows.append([
            '', 'TOTAL', '',
            f"{currency}{grand_premium:,.2f}",
            f"{currency}{grand_target:,.2f}",
            f"{currency}{grand_shares:,.2f}",
            f"{currency}{grand_total:,.2f}",
        ])

        col_widths = [2.2*cm, 5.5*cm, 4.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm]
        t = Table(table_rows, colWidths=col_widths, repeatRows=1)
        style = self._pdf_table_style(len(headers))
        style.add('FONTNAME',   (0, len(table_rows)-1), (-1, -1), 'Helvetica-Bold')
        style.add('BACKGROUND', (0, len(table_rows)-1), (-1, -1), ACCENT_COLOR)
        style.add('TEXTCOLOR',  (0, len(table_rows)-1), (-1, -1), WHITE)
        t.setStyle(style)
        story.append(t)
        doc.build(story)
        return out

    def savings_summary_excel(self, station_id: Optional[str] = None) -> str:
        q = """
            SELECT m.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name,
                   ROUND(SUM(CASE WHEN st.type_code='PREMIUM' THEN sa.current_balance ELSE 0 END),2) AS premium,
                   ROUND(SUM(CASE WHEN st.type_code='TARGET'  THEN sa.current_balance ELSE 0 END),2) AS target,
                   ROUND(SUM(CASE WHEN st.type_code='SHARES'  THEN sa.current_balance ELSE 0 END),2) AS shares,
                   ROUND(SUM(sa.current_balance),2) AS total
            FROM members m
            JOIN stations s ON m.station_id=s.station_id
            JOIN savings_accounts sa ON m.member_id=sa.member_id
            JOIN savings_types st ON sa.savings_type_id=st.savings_type_id
            WHERE m.is_active=1 AND m.is_deceased=0 AND sa.is_active=1
        """
        params = []
        if station_id:
            q += " AND m.station_id=?"; params.append(station_id)
        q += " GROUP BY m.member_id ORDER BY m.member_id"
        rows = self._q(q, tuple(params))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savings Summary"

        headers = ['Member ID', 'Full Name', 'Station',
                   'Premium', 'Target/Special', 'Shares', 'Total']
        next_row = self._xl_title(ws, 'Savings Summary Report',
                                   f'Members: {len(rows)}', len(headers))
        self._xl_header_row(ws, next_row, headers)
        next_row += 1

        for i, r in enumerate(rows):
            self._xl_data_row(ws, next_row, [
                r['member_id'], r['full_name'], r['station_name'],
                r['premium'] or 0, r['target'] or 0,
                r['shares'] or 0, r['total'] or 0
            ], alt=(i % 2 == 1))
            for col in range(4, 8):
                ws.cell(row=next_row, column=col).number_format = '#,##0.00'
            next_row += 1

        # Totals
        for j, val in enumerate(['', 'TOTAL', '',
                                  sum(r['premium'] or 0 for r in rows),
                                  sum(r['target'] or 0 for r in rows),
                                  sum(r['shares'] or 0 for r in rows),
                                  sum(r['total'] or 0 for r in rows)]):
            c = ws.cell(row=next_row, column=j+1, value=val)
            c.font   = Font(bold=True, color='FFFFFF')
            c.fill   = XL_SUBHDR_FILL
            c.border = XL_BORDER
            if j >= 3:
                c.number_format = '#,##0.00'

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        out = self._out('savings_summary', 'xlsx')
        wb.save(out)
        return out

    # -------------------------------------------------------------------------
    # 3. Loans Summary
    # -------------------------------------------------------------------------

    def loans_summary_pdf(self, station_id: Optional[str] = None,
                           status: str = 'Active') -> str:
        q = """
            SELECT l.loan_number, m.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name, lt.type_name,
                   l.principal_amount, l.interest_rate, l.total_amount,
                   l.amount_paid, l.balance_outstanding,
                   l.start_date, l.end_date, l.status
            FROM loans l
            JOIN members m ON l.member_id=m.member_id
            JOIN stations s ON l.station_id=s.station_id
            JOIN loan_types lt ON l.loan_type_id=lt.loan_type_id
            WHERE 1=1
        """
        params = []
        if station_id:
            q += " AND l.station_id=?"; params.append(station_id)
        if status != 'All':
            q += " AND l.status=?"; params.append(status)
        q += " ORDER BY m.member_id, l.start_date"
        loans    = self._q(q, tuple(params))
        currency = self._currency()

        total_principal   = sum(r['principal_amount'] or 0 for r in loans)
        total_outstanding = sum(r['balance_outstanding'] or 0 for r in loans)
        total_paid        = sum(r['amount_paid'] or 0 for r in loans)

        out    = self._out('loans_summary', 'pdf')
        doc    = SimpleDocTemplate(out, pagesize=landscape(A4),
                                   leftMargin=1*cm, rightMargin=1*cm,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = self._styles()
        story  = self._pdf_header(
            styles, f"Loans Summary — {status}",
            f"Loans: {len(loans)}  |  "
            f"Outstanding: {currency}{total_outstanding:,.2f}  |  "
            f"Collected: {currency}{total_paid:,.2f}"
        )

        headers = ['Loan No', 'Member ID', 'Name', 'Station', 'Type',
                   'Principal', 'Paid', 'Outstanding', 'Status']
        table_rows = [headers]
        for r in loans:
            table_rows.append([
                r['loan_number'], r['member_id'], r['full_name'],
                r['station_name'], r['type_name'],
                f"{currency}{r['principal_amount'] or 0:,.0f}",
                f"{currency}{r['amount_paid'] or 0:,.0f}",
                f"{currency}{r['balance_outstanding'] or 0:,.0f}",
                r['status']
            ])
        table_rows.append([
            '', '', 'TOTAL', '', '',
            f"{currency}{total_principal:,.0f}",
            f"{currency}{total_paid:,.0f}",
            f"{currency}{total_outstanding:,.0f}",
            ''
        ])

        col_widths = [3.5*cm, 2*cm, 4.5*cm, 3.5*cm, 3*cm,
                      3*cm, 3*cm, 3*cm, 2*cm]
        t = Table(table_rows, colWidths=col_widths, repeatRows=1)
        style = self._pdf_table_style(len(headers))
        style.add('FONTNAME',   (0, len(table_rows)-1), (-1, -1), 'Helvetica-Bold')
        style.add('BACKGROUND', (0, len(table_rows)-1), (-1, -1), ACCENT_COLOR)
        style.add('TEXTCOLOR',  (0, len(table_rows)-1), (-1, -1), WHITE)
        t.setStyle(style)
        story.append(t)
        doc.build(story)
        return out

    def loans_summary_excel(self, station_id: Optional[str] = None,
                             status: str = 'Active') -> str:
        q = """
            SELECT l.loan_number, m.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name, lt.type_name,
                   l.principal_amount, l.interest_amount, l.total_amount,
                   l.monthly_installment, l.duration_months,
                   l.amount_paid, l.balance_outstanding,
                   l.start_date, l.end_date, l.status
            FROM loans l
            JOIN members m ON l.member_id=m.member_id
            JOIN stations s ON l.station_id=s.station_id
            JOIN loan_types lt ON l.loan_type_id=lt.loan_type_id
            WHERE 1=1
        """
        params = []
        if station_id:
            q += " AND l.station_id=?"; params.append(station_id)
        if status != 'All':
            q += " AND l.status=?"; params.append(status)
        q += " ORDER BY m.member_id, l.start_date"
        loans = self._q(q, tuple(params))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Loans {status}"

        headers = ['Loan Number', 'Member ID', 'Full Name', 'Station', 'Type',
                   'Principal', 'Interest', 'Total', 'Monthly Install.',
                   'Duration (mo)', 'Amount Paid', 'Outstanding',
                   'Start Date', 'End Date', 'Status']
        next_row = self._xl_title(ws, f'Loans Summary — {status}',
                                   f'Total loans: {len(loans)}', len(headers))
        self._xl_header_row(ws, next_row, headers)
        next_row += 1

        num_cols = {5, 6, 7, 8, 10, 11}
        for i, r in enumerate(loans):
            self._xl_data_row(ws, next_row, [
                r['loan_number'], r['member_id'], r['full_name'],
                r['station_name'], r['type_name'],
                r['principal_amount'] or 0, r['interest_amount'] or 0,
                r['total_amount'] or 0, r['monthly_installment'] or 0,
                r['duration_months'] or 0,
                r['amount_paid'] or 0, r['balance_outstanding'] or 0,
                r['start_date'] or '', r['end_date'] or '', r['status']
            ], alt=(i % 2 == 1))
            for col in num_cols:
                ws.cell(row=next_row, column=col).number_format = '#,##0.00'
            next_row += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)

        out = self._out('loans_summary', 'xlsx')
        wb.save(out)
        return out

    # -------------------------------------------------------------------------
    # 4. Member Statement
    # -------------------------------------------------------------------------

    def member_statement_pdf(self, member_id: str) -> str:
        member = self._q("SELECT * FROM members WHERE member_id=?", (member_id,))
        if not member:
            raise ValueError(f"Member {member_id} not found")
        m        = member[0]
        currency = self._currency()
        full_name = f"{m['first_name']} {m.get('middle_name') or ''} {m['last_name']}".strip()

        station = self._q("SELECT station_name FROM stations WHERE station_id=?",
                          (m['station_id'],))
        station_name = station[0]['station_name'] if station else m['station_id']

        savings = self._q("""
            SELECT sa.account_number, st.type_name,
                   sa.current_balance, sa.total_deposits, sa.total_withdrawals,
                   sa.interest_earned
            FROM savings_accounts sa
            JOIN savings_types st ON sa.savings_type_id=st.savings_type_id
            WHERE sa.member_id=? AND sa.is_active=1
        """, (member_id,))

        loans = self._q("""
            SELECT l.loan_number, lt.type_name, l.principal_amount,
                   l.total_amount, l.amount_paid, l.balance_outstanding,
                   l.start_date, l.end_date, l.status
            FROM loans l
            JOIN loan_types lt ON l.loan_type_id=lt.loan_type_id
            WHERE l.member_id=?
            ORDER BY l.start_date DESC
        """, (member_id,))

        out    = self._out(f'statement_{member_id}', 'pdf')
        doc    = SimpleDocTemplate(out, pagesize=A4,
                                   leftMargin=2*cm, rightMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
        styles = self._styles()
        story  = self._pdf_header(styles, "Member Statement",
                                   f"{member_id} — {full_name}")

        # Member info
        story.append(Paragraph("Member Information", styles['SectionHdr']))
        info_data = [
            ['Member ID:', member_id,       'Station:',   station_name],
            ['Full Name:', full_name,        'Gender:',    m['gender'] or '—'],
            ['Date Joined:', m['date_joined'] or '—', 'Phone:', m['phone_number'] or '—'],
            ['Employee ID:', m['employee_id'] or '—', 'Grade:', m['grade_level'] or '—'],
        ]
        info_table = Table(info_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME',  (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE',  (0, 0), (-1, -1), 9),
            ('FONTNAME',  (0, 0), (0, -1),  'Helvetica-Bold'),
            ('FONTNAME',  (2, 0), (2, -1),  'Helvetica-Bold'),
            ('VALIGN',    (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',(0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 10))

        # Savings accounts
        if savings:
            story.append(Paragraph("Savings Accounts", styles['SectionHdr']))
            s_headers = ['Account No', 'Type', 'Balance', 'Total Deposits',
                         'Total Withdrawals', 'Interest Earned']
            s_rows    = [s_headers]
            for s in savings:
                s_rows.append([
                    s['account_number'], s['type_name'],
                    f"{currency}{s['current_balance'] or 0:,.2f}",
                    f"{currency}{s['total_deposits'] or 0:,.2f}",
                    f"{currency}{s['total_withdrawals'] or 0:,.2f}",
                    f"{currency}{s['interest_earned'] or 0:,.2f}",
                ])
            total_savings = sum(s['current_balance'] or 0 for s in savings)
            s_rows.append(['', 'TOTAL', f"{currency}{total_savings:,.2f}", '', '', ''])

            st = Table(s_rows, colWidths=[3.5*cm, 4*cm, 3*cm, 3*cm, 3*cm, 3*cm],
                       repeatRows=1)
            style = self._pdf_table_style(len(s_headers))
            style.add('FONTNAME',   (0, len(s_rows)-1), (-1, -1), 'Helvetica-Bold')
            style.add('BACKGROUND', (0, len(s_rows)-1), (-1, -1), ACCENT_COLOR)
            style.add('TEXTCOLOR',  (0, len(s_rows)-1), (-1, -1), WHITE)
            st.setStyle(style)
            story.append(st)
            story.append(Spacer(1, 10))

        # Loans
        if loans:
            story.append(Paragraph("Loan History", styles['SectionHdr']))
            l_headers = ['Loan No', 'Type', 'Principal', 'Total',
                         'Paid', 'Outstanding', 'Status']
            l_rows    = [l_headers]
            for l in loans:
                l_rows.append([
                    l['loan_number'], l['type_name'],
                    f"{currency}{l['principal_amount'] or 0:,.0f}",
                    f"{currency}{l['total_amount'] or 0:,.0f}",
                    f"{currency}{l['amount_paid'] or 0:,.0f}",
                    f"{currency}{l['balance_outstanding'] or 0:,.0f}",
                    l['status']
                ])
            lt = Table(l_rows, colWidths=[4.5*cm, 3.5*cm, 2.5*cm, 2.5*cm,
                                          2.5*cm, 2.5*cm, 2*cm], repeatRows=1)
            lt.setStyle(self._pdf_table_style(len(l_headers)))
            story.append(lt)

        doc.build(story)
        return out

    # -------------------------------------------------------------------------
    # 5. Transaction Report
    # -------------------------------------------------------------------------

    def transactions_report_pdf(self, start_date: str, end_date: str,
                                  member_id: Optional[str] = None,
                                  station_id: Optional[str] = None) -> str:
        q = """
            SELECT t.transaction_date, t.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name, t.transaction_type, t.account_type,
                   t.amount, t.is_credit, t.payment_method
            FROM transactions t
            LEFT JOIN members m ON t.member_id=m.member_id
            LEFT JOIN stations s ON t.station_id=s.station_id
            WHERE t.transaction_date >= ? AND t.transaction_date <= ?
        """
        params = [start_date, end_date]
        if member_id:
            q += " AND t.member_id=?"; params.append(member_id)
        if station_id:
            q += " AND t.station_id=?"; params.append(station_id)
        q += " ORDER BY t.transaction_date DESC, t.transaction_id DESC"
        txns     = self._q(q, tuple(params))
        currency = self._currency()

        total_credits = sum(r['amount'] or 0 for r in txns if r['is_credit'])
        total_debits  = sum(r['amount'] or 0 for r in txns if not r['is_credit'])

        out    = self._out('transactions', 'pdf')
        doc    = SimpleDocTemplate(out, pagesize=landscape(A4),
                                   leftMargin=1.5*cm, rightMargin=1.5*cm,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = self._styles()
        story  = self._pdf_header(
            styles, "Transaction Report",
            f"{start_date} to {end_date}  |  "
            f"Records: {len(txns)}  |  "
            f"Credits: {currency}{total_credits:,.2f}  |  "
            f"Debits: {currency}{total_debits:,.2f}"
        )

        headers   = ['Date', 'Member ID', 'Name', 'Station',
                     'Type', 'Account', 'Amount', 'Dr/Cr', 'Method']
        table_rows = [headers]
        for r in txns:
            table_rows.append([
                r['transaction_date'], r['member_id'], r['full_name'] or '—',
                r['station_name'] or '—', r['transaction_type'],
                r['account_type'],
                f"{currency}{r['amount'] or 0:,.2f}",
                'Cr' if r['is_credit'] else 'Dr',
                r['payment_method'] or '—'
            ])

        col_widths = [2.2*cm, 2*cm, 4.5*cm, 3.5*cm, 3.5*cm, 2.5*cm, 3*cm, 1.5*cm, 2.5*cm]
        t = Table(table_rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(self._pdf_table_style(len(headers)))
        story.append(t)
        doc.build(story)
        return out

    def transactions_report_excel(self, start_date: str, end_date: str,
                                    member_id: Optional[str] = None,
                                    station_id: Optional[str] = None) -> str:
        q = """
            SELECT t.transaction_date, t.member_id,
                   m.first_name || ' ' || COALESCE(m.middle_name || ' ','') || m.last_name AS full_name,
                   s.station_name, t.transaction_type, t.account_type,
                   t.amount, t.is_credit, t.payment_method,
                   t.cheque_number, t.receipt_number, t.description
            FROM transactions t
            LEFT JOIN members m ON t.member_id=m.member_id
            LEFT JOIN stations s ON t.station_id=s.station_id
            WHERE t.transaction_date >= ? AND t.transaction_date <= ?
        """
        params = [start_date, end_date]
        if member_id:
            q += " AND t.member_id=?"; params.append(member_id)
        if station_id:
            q += " AND t.station_id=?"; params.append(station_id)
        q += " ORDER BY t.transaction_date DESC, t.transaction_id DESC"
        txns = self._q(q, tuple(params))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ['Date', 'Member ID', 'Full Name', 'Station', 'Type',
                   'Account', 'Amount', 'Dr/Cr', 'Method',
                   'Cheque No', 'Receipt No', 'Description']
        next_row = self._xl_title(ws, 'Transaction Report',
                                   f'{start_date} to {end_date}  |  Records: {len(txns)}',
                                   len(headers))
        self._xl_header_row(ws, next_row, headers)
        next_row += 1

        for i, r in enumerate(txns):
            self._xl_data_row(ws, next_row, [
                r['transaction_date'], r['member_id'], r['full_name'] or '',
                r['station_name'] or '', r['transaction_type'],
                r['account_type'], r['amount'] or 0,
                'Cr' if r['is_credit'] else 'Dr',
                r['payment_method'] or '', r['cheque_number'] or '',
                r['receipt_number'] or '', r['description'] or ''
            ], alt=(i % 2 == 1))
            ws.cell(row=next_row, column=7).number_format = '#,##0.00'
            next_row += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        out = self._out('transactions', 'xlsx')
        wb.save(out)
        return out